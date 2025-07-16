import csv
import os
import sys
from datetime import datetime, date

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def str_to_date(date_str: str) -> date:
    """
    将形如'2024-12-17'的字符串转换成datetime.date对象
    :param date_str: 日期字符串。
    :return: date对象。
    """
    return datetime.strptime(date_str, '%Y-%m-%d').date()


def csv_to_list(csv_file_path: str) -> list:
    """
    将csv文件转换成列表，方便进一步处理。

    :param csv_file_path: csv文件的路径。
    :return: 转换后的列表。
    """
    with open(csv_file_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        return [[row[0], str_to_date(row[1]), int(row[2])] for row in reader]


def get_top_10(info_list: list) -> list:
    """
    从列表中获得前十浏览次数的招聘信息。

    :param info_list: 招聘信息列表。
    :return: 前十浏览次数的招聘信息的列表。
    """
    # 按照第3列（浏览次数）排序
    sorted_list = sorted(info_list, key=lambda x: x[2], reverse=True)
    # 返回前十
    return sorted_list[:10]


def list_to_xlsx(info: list, sheet: Worksheet, sheet_header: tuple | None = None) -> None:
    """
    将列表写入工作表。

    :param info: 待写入的列表。
    :param sheet: 待被写入的工作表。
    :param sheet_header: 可选的表头。
    :return: None
    """
    if sheet_header:
        sheet.append(sheet_header)

    index = 1
    for row in info:
        sheet.append([index] + row)
        index += 1


def get_both_top_10(info_list_1: list, info_list_2: list) -> list:
    """
    获取两个列表中都出现的、浏览次数前十的招聘信息。

    :param info_list_1: 招聘信息列表。
    :param info_list_2: 另一个招聘信息列表。
    :return: 前十浏览次数的招聘信息的列表。
    """
    temp = {}  # 存第一个列表的字典，键值对为（主题，行）
    result = []  # 结果

    for row in info_list_1:
        temp[row[0]] = row

    for row in info_list_2:
        # 利用hash原理，找到同时存在于两个文件的招聘信息
        if row[0] in temp:
            result.append(temp[row[0]] + [row[1], row[2], temp[row[0]][2] + row[2]])
    sorted_result = sorted(result, key=lambda x: x[5], reverse=True)
    return sorted_result[:10]


def adjust_col_width(sheet: Worksheet, col: str, wid: int) -> None:
    """
    调整列宽。

    :param sheet: 待调整的工作表。
    :param col: 待调整的列。（单个字母，如果不是，则不改变）
    :param wid: 调整后的宽度。
    :return: None
    """
    if col not in set('ABCDEFGHIJKLMNOPQRSTUVWXYZ'):
        return
    sheet.column_dimensions[col].width = wid


if __name__ == '__main__':
    # 爬取两个学校的信息，并将日志级别设为INFO
    print('正在爬取北邮信息...')
    os.system("scrapy crawl -L INFO BUPT")
    print('北邮信息爬取成功！')
    print('正在爬取西电信息...')
    os.system("scrapy crawl -L INFO XiDian")
    print('西电信息爬取成功！')

    print('正在将信息写入xlsx...')
    # 创建工作簿
    wb = openpyxl.Workbook()

    # 填入北邮信息
    ws = wb.active
    ws.title = '北邮'
    header = ('序号', '招聘主题', '发布日期', '浏览次数')
    bupt_list = csv_to_list('BUPT.csv')
    list_to_xlsx(bupt_list, ws, header)

    # 填入西电信息
    wb.create_sheet('西电')
    xidian_list = csv_to_list('XiDian.csv')
    list_to_xlsx(xidian_list, wb['西电'], header)

    # 获取并写入北邮招聘TOP10
    wb.create_sheet('北邮招聘TOP10')
    bupt_top_10 = get_top_10(bupt_list)
    list_to_xlsx(bupt_top_10, wb['北邮招聘TOP10'], header)

    # 获取并写入西电招聘TOP10
    wb.create_sheet('西电招聘TOP10')
    xidian_top_10 = get_top_10(xidian_list)
    list_to_xlsx(xidian_top_10, wb['西电招聘TOP10'], header)

    # 获取并写入共同TOP10
    wb.create_sheet('两校招聘TOP10')
    header = ('序号', '招聘信息', '北邮发布日期', '北邮浏览次数', '西电发布日期', '西电浏览次数', '浏览次数之和')
    both_top_10 = get_both_top_10(bupt_list, xidian_list)
    list_to_xlsx(both_top_10, wb['两校招聘TOP10'], header)

    # 调整列宽以便日期正常显示
    adjust_col_width(wb['北邮'], 'C', 15)
    adjust_col_width(wb['西电'], 'C', 15)
    adjust_col_width(wb['北邮招聘TOP10'], 'C', 15)
    adjust_col_width(wb['西电招聘TOP10'], 'C', 15)
    adjust_col_width(wb['两校招聘TOP10'], 'C', 15)
    adjust_col_width(wb['两校招聘TOP10'], 'E', 15)

    wb.save('就业信息汇总.xlsx')

    print('All Done!')
